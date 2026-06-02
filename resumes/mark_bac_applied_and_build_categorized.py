import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.csv")
XLSX_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.xlsx")
OUT_XLSX = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-tracker-categorized.xlsx")

APPLIED_DATE = "2026-06-03"
FOLLOW_UP_DATE = "2026-06-10"
GMAIL_ID = "19e89c01053d2189"


def normalize(value):
    return (value or "").strip()


with CSV_PATH.open(encoding="utf-8", newline="") as source:
    rows = list(csv.DictReader(source))
    fieldnames = list(rows[0].keys())

for row in rows:
    if normalize(row["Company / Agency"]) == "BAC Middle East":
        row["Status"] = "Applied"
        row["Date Applied"] = APPLIED_DATE
        row["Follow-up Date"] = FOLLOW_UP_DATE
        row["Recommended Action"] = "Follow up if no response"
        note = (
            "BAC-specific CV sent. To: recruit@bacme.com, submit@bacme.com. "
            f"CC: mabel@bacme.com. Gmail sent id: {GMAIL_ID}"
        )
        row["Notes"] = note if not row.get("Notes") else f"{row['Notes']} | {note}"

with CSV_PATH.open("w", encoding="utf-8", newline="") as output:
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

if XLSX_PATH.exists():
    workbook = load_workbook(XLSX_PATH)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    for row_idx in range(2, sheet.max_row + 1):
        company = sheet.cell(row=row_idx, column=headers.index("Company / Agency") + 1).value
        if normalize(company) == "BAC Middle East":
            values = {
                "Status": "Applied",
                "Date Applied": APPLIED_DATE,
                "Follow-up Date": FOLLOW_UP_DATE,
                "Recommended Action": "Follow up if no response",
            }
            for key, value in values.items():
                if key in headers:
                    sheet.cell(row=row_idx, column=headers.index(key) + 1).value = value
            if "Notes" in headers:
                note_cell = sheet.cell(row=row_idx, column=headers.index("Notes") + 1)
                note = f"BAC-specific CV sent. Gmail sent id: {GMAIL_ID}"
                note_cell.value = note if not note_cell.value else f"{note_cell.value} | {note}"
            break
    workbook.save(XLSX_PATH)


def write_sheet(workbook, title, sheet_rows):
    sheet = workbook.create_sheet(title)
    sheet.append(fieldnames)
    for row in sheet_rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    header_fill = PatternFill("solid", fgColor="242424")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for idx, field in enumerate(fieldnames, start=1):
        width = 18
        if field in {"Company / Agency", "Email", "Notes", "Opening Source URL(s)", "Recommended Action"}:
            width = 34 if field != "Opening Source URL(s)" else 56
        if field == "Matched Role(s)":
            width = 32
        sheet.column_dimensions[get_column_letter(idx)].width = width
    return sheet


wb = Workbook()
default = wb.active
wb.remove(default)

agency = [row for row in rows if normalize(row.get("Type")) == "Agency"]
direct = [row for row in rows if normalize(row.get("Type")) == "Direct Employer"]
linkedin = [row for row in rows if "linkedin" in normalize(row.get("Type")).lower() or "search" in normalize(row.get("Type")).lower()]
whatsapp = [row for row in rows if "whatsapp" in normalize(row.get("Email")).lower()]
applied = [row for row in rows if normalize(row.get("Status")).lower() == "applied"]
followups = [row for row in rows if normalize(row.get("Follow-up Date"))]
needs_email = [
    row
    for row in rows
    if "@" not in normalize(row.get("Email")) and "whatsapp" not in normalize(row.get("Email")).lower()
]

write_sheet(wb, "All", rows)
write_sheet(wb, "Agencies", agency)
write_sheet(wb, "Direct Employers", direct)
write_sheet(wb, "LinkedIn Search", linkedin)
write_sheet(wb, "WhatsApp Other", whatsapp)
write_sheet(wb, "Applied", applied)
write_sheet(wb, "Follow Ups", followups)
write_sheet(wb, "Needs Email Research", needs_email)

wb.save(OUT_XLSX)
print(OUT_XLSX)
