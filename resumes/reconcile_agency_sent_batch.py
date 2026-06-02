import csv
from pathlib import Path

from openpyxl import load_workbook

CSV_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.csv")
XLSX_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.xlsx")
APPLIED_DATE = "2026-06-03"
FOLLOW_UP_DATE = "2026-06-10"

SENT = {
    "ADMS UAE": "19e89945308caee8",
    "AKRS": "19e899494cedcd60",
    "ASRS": "19e8994d82744ae3",
    "Adecco": "19e89951d50b68b7",
    "Al Mansoor Group": "19e89956b48c3cad",
    "Al Nahiya": "19e8995bb9aefbbb",
    "Al Thawiya": "19e8995fe4f14b21",
    "Al Vakil": "19e8996472185e94",
    "Antal": "19e8996a624036db",
}

NOTE_PREFIX = "Agency outreach sent with UAE Product Designer CV"

with CSV_PATH.open(encoding="utf-8", newline="") as source:
    rows = list(csv.DictReader(source))
    fieldnames = list(rows[0].keys())

for row in rows:
    company = row["Company / Agency"]
    if company in SENT:
        row["Status"] = "Applied"
        row["Date Applied"] = APPLIED_DATE
        row["Follow-up Date"] = FOLLOW_UP_DATE
        note = f"{NOTE_PREFIX}. Gmail sent id: {SENT[company]}"
        existing = row.get("Notes", "")
        row["Notes"] = note if not existing else f"{existing} | {note}"
        row["Recommended Action"] = "Follow up if no response"

with CSV_PATH.open("w", encoding="utf-8", newline="") as output:
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

workbook = load_workbook(XLSX_PATH)
sheet = workbook.active
headers = [cell.value for cell in sheet[1]]

for row_idx in range(2, sheet.max_row + 1):
    company = sheet.cell(row=row_idx, column=headers.index("Company / Agency") + 1).value
    if company in SENT:
        sheet.cell(row=row_idx, column=headers.index("Status") + 1).value = "Applied"
        sheet.cell(row=row_idx, column=headers.index("Date Applied") + 1).value = APPLIED_DATE
        sheet.cell(row=row_idx, column=headers.index("Follow-up Date") + 1).value = FOLLOW_UP_DATE
        notes_cell = sheet.cell(row=row_idx, column=headers.index("Notes") + 1)
        note = f"{NOTE_PREFIX}. Gmail sent id: {SENT[company]}"
        notes_cell.value = note if not notes_cell.value else f"{notes_cell.value} | {note}"
        sheet.cell(row=row_idx, column=headers.index("Recommended Action") + 1).value = "Follow up if no response"

workbook.save(XLSX_PATH)
print(f"Reconciled {len(SENT)} agency applications")
