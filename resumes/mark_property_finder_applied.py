import csv

from openpyxl import load_workbook

CSV_PATH = r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.csv"
XLSX_PATH = r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.xlsx"
GMAIL_ID = "19e896596bf0a734"

with open(CSV_PATH, encoding="utf-8", newline="") as source:
    rows = list(csv.DictReader(source))
    fieldnames = list(rows[0].keys())

for row in rows:
    if row["Company / Agency"] == "Property Finder":
        row["Status"] = "Applied"
        row["Date Applied"] = "2026-06-02"
        row["Follow-up Date"] = "2026-06-09"
        row["Notes"] = f"{row.get('Notes', '')} | Gmail sent id: {GMAIL_ID}".strip()

with open(CSV_PATH, "w", encoding="utf-8", newline="") as output:
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

workbook = load_workbook(XLSX_PATH)
sheet = workbook.active
headers = [cell.value for cell in sheet[1]]

for row_idx in range(2, sheet.max_row + 1):
    if sheet.cell(row=row_idx, column=headers.index("Company / Agency") + 1).value == "Property Finder":
        sheet.cell(row=row_idx, column=headers.index("Status") + 1).value = "Applied"
        sheet.cell(row=row_idx, column=headers.index("Date Applied") + 1).value = "2026-06-02"
        sheet.cell(row=row_idx, column=headers.index("Follow-up Date") + 1).value = "2026-06-09"
        notes_cell = sheet.cell(row=row_idx, column=headers.index("Notes") + 1)
        notes_cell.value = f"{notes_cell.value or ''} | Gmail sent id: {GMAIL_ID}".strip()
        break

workbook.save(XLSX_PATH)
print("Property Finder marked Applied")
