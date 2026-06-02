import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.csv"
OUT = r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.xlsx"

with open(SRC, encoding="utf-8", newline="") as source:
    rows = list(csv.reader(source))

workbook = Workbook()
sheet = workbook.active
sheet.title = "Opening Scan"

for row in rows:
    sheet.append(row)

header_fill = PatternFill("solid", fgColor="242424")
header_font = Font(color="FFFFFF", bold=True)
for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font

status_col = rows[0].index("Opening Scan Status") + 1
confidence_col = rows[0].index("Confidence") + 1
for row in range(2, sheet.max_row + 1):
    status = sheet.cell(row=row, column=status_col).value or ""
    confidence = sheet.cell(row=row, column=confidence_col).value or ""
    if "Opening" in status:
        fill = PatternFill("solid", fgColor="D9EAD3")
    elif "No Relevant" in status or "No Clear" in status:
        fill = PatternFill("solid", fgColor="FCE5CD")
    else:
        fill = PatternFill("solid", fgColor="FFF2CC")
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=row, column=col).fill = fill
    if confidence == "High":
        sheet.cell(row=row, column=confidence_col).font = Font(bold=True, color="38761D")

widths = {
    "A": 28,
    "B": 34,
    "C": 18,
    "D": 16,
    "H": 28,
    "I": 42,
    "J": 70,
    "K": 14,
    "L": 34,
}
for col_idx in range(1, sheet.max_column + 1):
    letter = get_column_letter(col_idx)
    sheet.column_dimensions[letter].width = widths.get(letter, 18)

sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions
workbook.save(OUT)
print(OUT)
