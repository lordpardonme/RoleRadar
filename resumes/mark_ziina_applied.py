import csv
from pathlib import Path

from openpyxl import load_workbook

CSV_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.csv")
XLSX_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.xlsx")
GMAIL_ID = "19e898b575624043"

with CSV_PATH.open(encoding="utf-8", newline="") as source:
    rows = list(csv.DictReader(source))
    fieldnames = list(rows[0].keys())

row = {
    field: ""
    for field in fieldnames
}
row.update(
    {
        "Company / Agency": "Ziina",
        "Email": "anton.badashov@ziina.com; jocelyn.meyer@ziina.com",
        "Type": "Direct Employer",
        "Status": "Applied",
        "Date Applied": "2026-06-02",
        "Follow-up Date": "2026-06-09",
        "Notes": f"Sent tailored Ziina fintech CV. Gmail sent id: {GMAIL_ID}",
        "Opening Scan Status": "Opening Found",
        "Matched Role(s)": "Senior Product Designer",
        "Opening Source URL(s)": "https://job-boards.greenhouse.io/ziina/jobs/4631469101",
        "Confidence": "High",
        "Recommended Action": "Follow up in one week if no reply",
        "Scan Date": "2026-06-02",
    }
)

found = False
for existing in rows:
    if existing["Company / Agency"].strip().lower() == "ziina":
        existing.update(row)
        found = True
        break
if not found:
    rows.append(row)

with CSV_PATH.open("w", encoding="utf-8", newline="") as output:
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

workbook = load_workbook(XLSX_PATH)
sheet = workbook.active
headers = [cell.value for cell in sheet[1]]
company_col = headers.index("Company / Agency") + 1

target_row = None
for idx in range(2, sheet.max_row + 1):
    value = sheet.cell(row=idx, column=company_col).value
    if isinstance(value, str) and value.strip().lower() == "ziina":
        target_row = idx
        break

if target_row is None:
    target_row = sheet.max_row + 1

for header, value in row.items():
    if header in headers:
        sheet.cell(row=target_row, column=headers.index(header) + 1).value = value

workbook.save(XLSX_PATH)
print("Ziina marked Applied")
