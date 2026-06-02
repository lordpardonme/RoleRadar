import csv

from openpyxl import load_workbook

CSV_PATH = r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.csv"
XLSX_PATH = r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.xlsx"

updates = {
    "Status": "Ready to Apply",
    "Notes": (
        "ATS 86/100. Tailored resume: resumes/property-finder-product-designer-resume.md. "
        "Report: resumes/property-finder-product-designer-ats-report.md. "
        "Figma: https://www.figma.com/design/or9EJUqIuF1yN5SEJ15LtP"
    ),
    "Recommended Action": "Apply now with tailored resume and email",
}

with open(CSV_PATH, encoding="utf-8", newline="") as source:
    rows = list(csv.DictReader(source))
    fieldnames = source.seek(0) or list(rows[0].keys())

for row in rows:
    if row["Company / Agency"] == "Property Finder":
        row.update(updates)

with open(CSV_PATH, "w", encoding="utf-8", newline="") as output:
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)

workbook = load_workbook(XLSX_PATH)
sheet = workbook.active
headers = [cell.value for cell in sheet[1]]
company_col = headers.index("Company / Agency") + 1
for row_idx in range(2, sheet.max_row + 1):
    if sheet.cell(row=row_idx, column=company_col).value == "Property Finder":
        for header, value in updates.items():
            col_idx = headers.index(header) + 1
            sheet.cell(row=row_idx, column=col_idx).value = value
        break
workbook.save(XLSX_PATH)

print("Property Finder marked Ready to Apply")
