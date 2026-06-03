import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.csv")
XLSX_PATH = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-opening-scan.xlsx")
OUT_XLSX = Path(r"C:\Users\mohdh\Documents\Job Hunt\resumes\uae-job-tracker-categorized.xlsx")

DATE = "2026-06-03"
FOLLOW_UP = "2026-06-10"

RESENDS = {
    "ADMS UAE": "19e89cf5ce9cc34b",
    "AKRS": "19e89cfc2a56acd0",
    "ASRS": "19e89d0228492b70",
    "Adecco": "19e89d089904515a",
    "Al Mansoor Group": "19e89d0ead50a17f",
    "Al Nahiya": "19e89d146678c6a2",
    "Al Thawiya": "19e89d1b128adb5c",
    "Al Vakil": "19e89d20d8adcac1",
    "Antal": "19e89d28a4d760b1",
}


def add_note(existing, note):
    return note if not existing else f"{existing} | {note}"


with CSV_PATH.open(encoding="utf-8", newline="") as source:
    rows = list(csv.DictReader(source))
    fieldnames = list(rows[0].keys())

for row in rows:
    company = row["Company / Agency"]
    if company in RESENDS:
        row["Status"] = "Applied"
        row["Date Applied"] = DATE
        row["Follow-up Date"] = FOLLOW_UP
        row["Recommended Action"] = "Follow up if no response"
        row["Notes"] = add_note(
            row.get("Notes", ""),
            f"Duplicate resend approved by user; updated/tighter agency CV sent. Gmail sent id: {RESENDS[company]}",
        )

with CSV_PATH.open("w", encoding="utf-8", newline="") as output:
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

workbook = load_workbook(XLSX_PATH)
sheet = workbook.active
headers = [cell.value for cell in sheet[1]]

for row_idx in range(2, sheet.max_row + 1):
    company = sheet.cell(row=row_idx, column=headers.index("Company / Agency") + 1).value
    if company in RESENDS:
        sheet.cell(row=row_idx, column=headers.index("Status") + 1).value = "Applied"
        sheet.cell(row=row_idx, column=headers.index("Date Applied") + 1).value = DATE
        sheet.cell(row=row_idx, column=headers.index("Follow-up Date") + 1).value = FOLLOW_UP
        sheet.cell(row=row_idx, column=headers.index("Recommended Action") + 1).value = "Follow up if no response"
        notes_cell = sheet.cell(row=row_idx, column=headers.index("Notes") + 1)
        notes_cell.value = add_note(
            notes_cell.value or "",
            f"Duplicate resend approved by user; Gmail sent id: {RESENDS[company]}",
        )
workbook.save(XLSX_PATH)


def write_sheet(wb, title, sheet_rows):
    ws = wb.create_sheet(title)
    ws.append(fieldnames)
    for row in sheet_rows:
        ws.append([row.get(field, "") for field in fieldnames])
    header_fill = PatternFill("solid", fgColor="242424")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fieldnames, start=1):
        width = 18
        if field in {"Company / Agency", "Email", "Notes", "Opening Source URL(s)", "Recommended Action"}:
            width = 34 if field != "Opening Source URL(s)" else 56
        if field == "Matched Role(s)":
            width = 32
        ws.column_dimensions[get_column_letter(idx)].width = width


def norm(value):
    return (value or "").strip()


wb = Workbook()
wb.remove(wb.active)
write_sheet(wb, "All", rows)
write_sheet(wb, "Agencies", [r for r in rows if norm(r.get("Type")) == "Agency"])
write_sheet(wb, "Direct Employers", [r for r in rows if norm(r.get("Type")) == "Direct Employer"])
write_sheet(wb, "LinkedIn Search", [r for r in rows if "linkedin" in norm(r.get("Type")).lower() or "search" in norm(r.get("Type")).lower()])
write_sheet(wb, "WhatsApp Other", [r for r in rows if "whatsapp" in norm(r.get("Email")).lower()])
write_sheet(wb, "Applied", [r for r in rows if norm(r.get("Status")).lower() == "applied"])
write_sheet(wb, "Follow Ups", [r for r in rows if norm(r.get("Follow-up Date"))])
write_sheet(wb, "Needs Email Research", [r for r in rows if "@" not in norm(r.get("Email")) and "whatsapp" not in norm(r.get("Email")).lower()])
wb.save(OUT_XLSX)

print("Duplicate resend tracker updated")
